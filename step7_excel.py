#!/usr/bin/env python3

#from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, NamedStyle, numbers
from lib.json_io import load_json, save_json
from lib.string import capitalize
import json
import os
import pandas as pd
import sys

def prepare(df, keep_words):
    df['Transaction ID'] = df['transactionId'].astype(str)
    df['Date V/R'] = pd.to_datetime(df['valueDate'])
    df['Débiteur'] = df['debtorName'].apply(lambda x: capitalize(x, keep_words) if isinstance(x, str) else x)
    df['Compte D'] = df['debtorAccount'].apply(lambda x: x['iban'] if isinstance(x, dict) and 'iban' in x else None)
    df['Créditeur'] = df['creditorName'].apply(lambda x: capitalize(x, keep_words) if isinstance(x, str) else x)
    df['Compte C'] = df['creditorAccount'].apply(lambda x: x['iban'] if isinstance(x, dict) and 'iban' in x else None)
    df['Quoi'] = df['additionalInformation']
    df['Montant'] = df['transactionAmount'].apply(lambda x: float(x['amount']))
    df['Communication'] = df['additionalInformation']
    df['Sommaire'] = df['remittanceInformationUnstructured'].str.strip()
    return df[['Transaction ID', 'Date V/R', 'Débiteur', 'Compte D', 'Créditeur', 'Compte C', 'Quoi', 'Montant', 'Communication', 'Sommaire']]

def create(excel_file, sheet_name, df):
    df.to_excel(excel_file, sheet_name=sheet_name, index=False)
    print("Created new Excel file with transaction data")
    sys.exit(0)

def append(excel_file, sheet_name, df):
    columns = {
        'Transaction ID': 1,    # A
        #'Mois': 2,             # B
        'Date V/R': 3,          # C
        #'D Buy/Fac': 4,        # D
        #'Type': 5,             # E
        #'Mode': 6,             # F
        'Débiteur': 7,          # G
        'Compte D': 8,          # H
        'Créditeur': 9,         # I
        'Compte C': 10,         # J
        #'Shop': 11,            # K
        'Quoi': 12,             # L
        'Communication': 13,    # M
        'Montant': 14,          # N
        #'Total': 15,           # O
        #'Commentaire':16       # P
        'Sommaire': 17          # Q
    }

    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    # Find index of that ID in the DataFrame
    last_known_id = ws[ws.max_row][0].value
    if last_known_id in df['Transaction ID'].values:
        last_index = df.index[df['Transaction ID'] == last_known_id][0]
        df_to_append = df.iloc[last_index + 1:]
    else:
        df_to_append = df # If not found, assume entire DataFrame is new

    if df_to_append.shape[0] == 0: 
        print('Nothing to append; Excel file already up-to-date. Exiting~')
        sys.exit(0)

    # Write new rows
    print(df_to_append.shape[0], 'new rows to append')
    row_start = ws.max_row + 1
    df_to_append = df_to_append.fillna('-')
    for i, (_, row) in enumerate(df_to_append.iterrows(), start=row_start):
        for col_name, value in row.items():
            j = columns[col_name]
            ws.cell(row=i, column=j, value=value)

    wb.save(excel_file)

    print("Appended transaction data to existing Excel file")
    return row_start

def get_mode(sommaire):
    if not isinstance(sommaire, str): return None

    s = sommaire.strip().upper() # Probably unnecessary
    if s.startswith('VERSEMENT'): return 'In'
    elif s.startswith('ORDRE PERMANENT'): return 'Ordre'
    elif s.startswith('BANCONTACT ACHAT'): return 'Carte'
    elif s.startswith('VIREMENT INSTANTANE'): return 'PC'
    elif s.startswith('REMBOURSEMENT'): return 'Remboursement'
    elif s.startswith('PARTICIPATION AUX FRAIS DE TENUE DE VOTRE COMPTE') or s.startswith('COUT GESTION CARTE DE DEBIT'): return 'Banque'
    else: return 'UNKNOWN'

def format_bank_account(s: str) -> str:
    """
    Format a bank/IBAN string into 4-character groups separated by spaces.
    Non-space characters are kept; spaces are stripped before grouping.
    """
    clean = ''.join(s.split()).upper()
    return ' '.join(clean[i:i+4] for i in range(0, len(clean), 4))

def format(excel_file, sheet_name, row_start=0):
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    font = Font(name='Verdana', size=8)
    small_font = Font(name='Verdana', size=7)
    smaller_font = Font(name='Verdana', size=6)
    smallest_font = Font(name='Verdana', size=5)
    bold_font = Font(name='Verdana', size=8, bold=True)
    for row in ws.iter_rows(min_row=row_start, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = font

    for row in ws.iter_rows(min_row=row_start, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row: # Transaction ID
            cell.font = smaller_font
    for i, row in enumerate(ws.iter_rows(min_row=row_start, max_row=ws.max_row, min_col=2, max_col=2)):
        for cell in row: # Mois
            cell.font = small_font
            #cell.value = '=PROPER(TEXT(INDIRECT("C" & ROW()), "[$-40C]mmmm"))'
            cell.value = f'=PROPER(TEXT(C{row_start + i},"[$-40C]mmmm"))'
    for row in ws.iter_rows(min_row=row_start, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row: # Date V/R
            cell.number_format = 'yyyy.mm.dd'

    for i, row in enumerate(ws.iter_rows(min_row=row_start, max_row=ws.max_row, min_col=6, max_col=6)):
        for cell in row: # Mode
            cell.value = get_mode(ws.cell(row=row_start + i, column=17).value)

    for row in ws.iter_rows(min_row=row_start, max_row=ws.max_row, min_col=7, max_col=7):
        for cell in row: # Débiteur
            cell.alignment = Alignment(horizontal='left')
            cell.font = small_font if len(cell.value) > 26 else font
    for row in ws.iter_rows(min_row=row_start, max_row=ws.max_row, min_col=8, max_col=8):
        for cell in row: # Compte D
            cell.font = smallest_font
            cell.value = format_bank_account(cell.value) if isinstance(cell.value, str) else cell.value
    for row in ws.iter_rows(min_row=row_start, max_row=ws.max_row, min_col=9, max_col=9):
        for cell in row: # Créditeur
            cell.alignment = Alignment(horizontal='left')
            cell.font = small_font if len(cell.value) > 26 else font
    for row in ws.iter_rows(min_row=row_start, max_row=ws.max_row, min_col=10, max_col=10):
        for cell in row: # Compte C
            cell.font = smallest_font
            cell.value = format_bank_account(cell.value) if isinstance(cell.value, str) else cell.value
    for row in ws.iter_rows(min_row=row_start, max_row=ws.max_row, min_col=12, max_col=12):
        for cell in row: # Quoi
            cell.alignment = Alignment(horizontal='left')
            cell.font = small_font if len(cell.value) > 46 else font
    for row in ws.iter_rows(min_row=row_start, max_row=ws.max_row, min_col=13, max_col=13):
        for cell in row: # Communication
            cell.alignment = Alignment(horizontal='left')
            cell.font = smaller_font
    for row in ws.iter_rows(min_row=row_start, max_row=ws.max_row, min_col=14, max_col=14):
        for cell in row: # Montant
            cell.alignment = Alignment(horizontal='right')
            cell.font = small_font
            cell.number_format = '0.00'
    for i, row in enumerate(ws.iter_rows(min_row=row_start, max_row=ws.max_row, min_col=15, max_col=15)):
        for cell in row: # Total
            cell.alignment = Alignment(horizontal='right')
            cell.font = bold_font
            cell.number_format = '0.00'
            # cell.value = '=INDIRECT("N" & ROW())'
            cell.value = f'=N{row_start + i}'
    for row in ws.iter_rows(min_row=row_start, max_row=ws.max_row, min_col=16, max_col=16):
        for cell in row: # Commentaire
            cell.alignment = Alignment(horizontal='left')
    for row in ws.iter_rows(min_row=row_start, max_row=ws.max_row, min_col=17, max_col=17):
        for cell in row: # Sommaire
            cell.alignment = Alignment(horizontal='left')
            cell.font = smaller_font

    wb.save(excel_file)

def main(transactions_file, excel_file, sheet_name='Compte'):
    if not os.path.exists(transactions_file):
        print('Transactions file is missing. Aborting.')
        sys.exit(1)

    transactions = load_json(transactions_file)
    df = pd.DataFrame(transactions['transactions']['booked'])
    # df[::-1] inverts rows & reset_index makes it so index isn't reversed as well (making it pointless)
    filtered = prepare(df[::-1].reset_index(drop=True), { 'ASBL', 'G.B.R.S.' })   

    if not os.path.exists(excel_file):
        #create(excel_file, sheet_name, filtered) # Recreate
        print('The Excel file must exist prior to using this script')
        sys.exit(2)

    row_start = append(excel_file, sheet_name, filtered)
    format(excel_file, sheet_name, row_start)
    #os.system('powershell -Command "Unblock-File -Path \'output.xlsx\'"') 



if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: excel.py <transactions_file> <excel_file> (<sheet_name>)")
        sys.exit(1)

    transactions_file = sys.argv[1]
    excel_file = sys.argv[2]
    sheet_name = sys.argv[3] if len(sys.argv) > 3 else 'Compte'
    main(transactions_file, excel_file, sheet_name)

